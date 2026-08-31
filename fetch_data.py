#!/usr/bin/env python3
"""
Data pipeline for the NDQ.AX Pulse dashboard.

Fetches REAL data from:
  1. Yahoo Finance  - NDQ.AX, HNDQ.AX, TSLA, SPCX daily prices (adjusted close = total return)
  2. AustralianSuper official API - daily crediting-rate returns for every investment option (11 yrs)
  3. AustralianSuper official adviser Excel - published 1/3/5/10 yr returns (independent verification)
  4. Curated press-reported SpaceX private valuations (funding rounds / tender offers / IPO)
     - sources: CNBC, CNN, Forbes, Bloomberg, BBC, The Times, Fortune (see VALUATIONS below)

Output: data/dashboard.js  (window.DASHBOARD_DATA = {...})

Re-run any time to refresh:  .venv/bin/python fetch_data.py
"""

import json
import math
import time
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from curl_cffi import requests as cr

ROOT = Path(__file__).resolve().parent
RAW = ROOT / "data" / "raw"
RAW.mkdir(parents=True, exist_ok=True)

SYD = ZoneInfo("Australia/Sydney")
P1 = 1420070400  # 1 Jan 2015 UTC

# --------------------------------------------------------------------------- #
# HTTP helper                                                                 #
# --------------------------------------------------------------------------- #

def get_with_retry(url: str, attempts: int = 5, timeout: int = 110):
    """GET with exponential backoff — remote APIs occasionally answer 5xx."""
    r = None
    for i in range(attempts):
        try:
            r = cr.get(url, impersonate="chrome", timeout=timeout)
            if r.status_code == 200:
                return r
            wait = min(2 ** i * 5, 60)
            print(f"  HTTP {r.status_code} — retrying in {wait}s ({i+1}/{attempts})")
        except Exception as e:
            wait = min(2 ** i * 5, 60)
            print(f"  {e} — retrying in {wait}s ({i+1}/{attempts})")
        time.sleep(wait)
    r.raise_for_status()
    return r

# --------------------------------------------------------------------------- #
# 1. Yahoo Finance                                                            #
# --------------------------------------------------------------------------- #

def fetch_yahoo(symbol: str, period1: int = P1) -> pd.DataFrame:
    """Daily close + adjusted close from Yahoo Finance."""
    url = (f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
           f"?period1={period1}&period2={int(time.time()) + 86400}"
           f"&interval=1d&events=div%2Csplit")
    r = get_with_retry(url, timeout=60)
    payload = r.json()
    (RAW / f"{symbol}.json").write_text(json.dumps(payload))

    res = payload["chart"]["result"][0]
    ts = res["timestamp"]
    q = res["indicators"]["quote"][0]
    adj = res["indicators"]["adjclose"][0]["adjclose"]
    dates = [datetime.fromtimestamp(t, timezone.utc).astimezone(SYD).date() for t in ts]
    df = pd.DataFrame({"date": pd.to_datetime(dates),
                       "close": q["close"],
                       "adjclose": adj}).dropna()
    df = df.drop_duplicates(subset="date").sort_values("date").reset_index(drop=True)
    print(f"  {symbol}: {len(df)} rows  {df.date.min().date()} -> {df.date.max().date()}")
    return df

# --------------------------------------------------------------------------- #
# 2. AustralianSuper daily crediting rates (cumulative from May 2015)         #
# --------------------------------------------------------------------------- #

ASU_API = ("https://www.australiansuper.com/api/graphs/dailyrates/graph"
           "?superType=Super&start=26/05/2015&end={end}&cumulative=true"
           "&truncateDecimalPlaces=true")

def fetch_australiansuper() -> pd.DataFrame:
    end = datetime.now(SYD).strftime("%d/%m/%Y")
    r = get_with_retry(ASU_API.format(end=end))
    payload = r.json()
    (RAW / "australiansuper_daily.json").write_text(json.dumps(payload))

    groups = payload["drawing"][0]["products"]
    frames = {}
    for p in groups:
        rows = [(datetime.strptime(g["day"], "%d %b %Y").date(), float(g["rate"]))
                for g in p["graph"]]
        s = pd.DataFrame(rows, columns=["date", p["title"]])
        s["date"] = pd.to_datetime(s["date"])
        frames[p["title"]] = s.set_index("date")
    df = frames[list(frames)[0]].join([frames[k] for k in list(frames)[1:]], how="outer").sort_index()
    df = df.ffill()
    print(f"  AustralianSuper: {len(df)} rows  {df.index.min().date()} -> {df.index.max().date()}")
    return df

# --------------------------------------------------------------------------- #
# 3. AustralianSuper official published returns (adviser Excel)               #
# --------------------------------------------------------------------------- #

ASU_XLSX = ("https://www.australiansuper.com/-/media/australian-super/files/campaigns/"
            "adviser-resources/performance-data/super-and-retirement-excel.xlsx")

def fetch_official_table() -> dict:
    import openpyxl
    r = get_with_retry(ASU_XLSX, timeout=60)
    path = RAW / "australiansuper_official.xlsx"
    path.write_bytes(r.content)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Superannuation"]
    as_of = ws.cell(row=1, column=3).value
    out = {}
    for row in ws.iter_rows(min_row=9, values_only=True):
        name = row[2]
        if name in ("Australian Shares", "International Shares"):
            out[name] = {"1Y": row[8] * 100, "3Y": row[10] * 100,
                         "5Y": row[11] * 100, "10Y": row[13] * 100,
                         "inception": str(row[17])[:10]}
    return {"asOf": str(as_of)[:10], "options": out}

# --------------------------------------------------------------------------- #
# 4. SpaceX private-market valuation history (press-reported, curated)        #
# --------------------------------------------------------------------------- #
# SpaceX was private until its NASDAQ IPO (SPCX) on 12 June 2026. These are
# the funding-round / tender-offer valuations reported by major outlets.
# A few early dates are approximate to the month (noted in methodology).

SPACEX_VALUATIONS = [
    # date          $B      event                                             source
    ("2012-02-01",    1.3,  "Early private-market estimate",                  "PrivCo/CNBC"),
    ("2012-05-25",    2.4,  "Re-rate after first ISS docking ($20/share)",    "PrivCo"),
    ("2015-01-20",   12.0,  "Google & Fidelity invest $1B",                   "Reuters/CNBC"),
    ("2017-07-27",   21.2,  "Funding round (~$351M raised)",                  "CNBC"),
    ("2018-04-05",   27.0,  "2018 funding round",                             "Wikipedia/CNBC"),
    ("2019-05-31",   33.3,  "Starlink fundraising round",                     "CNBC"),
    ("2020-02-21",   36.0,  "$250M raise",                                    "CNBC"),
    ("2020-08-18",   46.0,  "$1.9B raise — largest to date",                  "CNN"),
    ("2021-02-16",   74.0,  "$850M raise @ $419.99/share",                    "CNBC"),
    ("2021-10-08",  100.0,  "Secondary share sale",                           "CNBC"),
    ("2022-08-01",  127.0,  "2022 funding round",                             "Wikipedia"),
    ("2023-01-02",  137.0,  "$750M raise (a16z)",                             "CNBC"),
    ("2023-12-13",  180.0,  "Tender offer @ $97/share",                       "CNBC"),
    ("2024-07-01",  210.0,  "Tender offer @ ~$112/share",                     "Bloomberg"),
    ("2024-12-03",  350.0,  "Tender offer — most valuable startup",           "Forbes"),
    ("2025-07-08",  400.0,  "Insider share sale",                             "Bloomberg"),
    ("2025-12-22",  800.0,  "Late-2025 tender talks",                         "Forbes"),
    ("2026-03-02", 1000.0,  "xAI merger values SpaceX at ~$1T",               "Wikipedia"),
    ("2026-06-11", 1800.0,  "IPO priced @ $135/share — raised $85.7B",        "BBC"),
]

# SPCX IPO facts (BBC, 15 Jun 2026): priced at $135/share, $1.8T valuation,
# $75B raised + $10.7B greenshoe = $85.7B total — largest IPO in history.
SPCX_IPO_PRICE = 135.0
SPCX_IPO_VALUATION_B = 1800.0
SPCX_IMPLIED_SHARES_B = SPCX_IPO_VALUATION_B / SPCX_IPO_PRICE   # ~13.33B shares

# --------------------------------------------------------------------------- #
# Processing helpers                                                          #
# --------------------------------------------------------------------------- #

def weekly_last(df: pd.DataFrame, col: str) -> pd.DataFrame:
    """Last observation per calendar week (keeps the real observation date)."""
    d = df[["date", col]].copy()
    d["week"] = d["date"].dt.to_period("W-SUN")
    out = d.groupby("week").tail(1).reset_index(drop=True)
    return out[["date", col]]

def trailing_pa(index_series: pd.Series, years: float, asof=None) -> float | None:
    """Annualised trailing return from a rebased index series (date-indexed)."""
    s = index_series.dropna()
    if len(s) < 2:
        return None
    end = asof or s.index[-1]
    start_target = end - pd.DateOffset(years=int(years)) if years == int(years) else end - pd.DateOffset(days=int(years*365.25))
    past = s[s.index <= start_target]
    if past.empty:
        return None
    v0, v1 = past.iloc[-1], s.asof(end)
    actual_years = (end - past.index[-1]).days / 365.25
    if actual_years < years * 0.95 or v0 <= 0:
        return None
    return ((v1 / v0) ** (1 / actual_years) - 1) * 100

def rebase(series: pd.Series, value: float = 100.0) -> pd.Series:
    s = series.dropna()
    return s / s.iloc[0] * value

def r2(x):
    return None if x is None or (isinstance(x, float) and (math.isnan(x) or math.isinf(x))) else round(float(x), 2)

def analyze_ma200(df: pd.DataFrame) -> dict:
    """Weekly close, 200-week SMA, contiguous below-MA buy zones, and status KPIs."""
    w = weekly_last(df, "close")
    w["ma200"] = w["close"].rolling(200, min_periods=200).mean()
    w["below"] = w["ma200"].notna() & (w["close"] < w["ma200"])

    zones, cur = [], None
    for _, row in w.iterrows():
        if row["below"]:
            pct = (row["close"] / row["ma200"] - 1) * 100
            if cur is None:
                cur = {"start": row["date"], "end": row["date"], "weeks": 1, "maxBelow": pct}
            else:
                cur["end"] = row["date"]; cur["weeks"] += 1
                cur["maxBelow"] = min(cur["maxBelow"], pct)
        elif cur is not None:
            zones.append(cur); cur = None
    if cur is not None:
        zones.append(cur)

    latest = df.iloc[-1]
    latest_ma = w["ma200"].dropna().iloc[-1] if w["ma200"].notna().any() else None

    return {
        "weekly": {
            "dates": [d.strftime("%Y-%m-%d") for d in w["date"]],
            "close": [r2(v) for v in w["close"]],
            "ma200": [r2(v) if not pd.isna(v) else None for v in w["ma200"]],
            "below": [bool(b) for b in w["below"]],
        },
        "zones": [{
            "start": z["start"].strftime("%Y-%m-%d"),
            "end": z["end"].strftime("%Y-%m-%d"),
            "weeks": z["weeks"],
            "maxBelowPct": r2(z["maxBelow"]),
        } for z in zones],
        "kpi": {
            "price": r2(latest["close"]),
            "date": latest["date"].strftime("%Y-%m-%d"),
            "ma200": r2(latest_ma) if latest_ma is not None else None,
            "vsMaPct": r2((latest["close"] / latest_ma - 1) * 100) if latest_ma else None,
            "inBuyZone": bool(latest_ma is not None and latest["close"] < latest_ma),
            "weeksInBuyZone": int(w["below"].sum()),
            "weeksWithMa": int(w["ma200"].notna().sum()),
            "buyZoneCount": len(zones),
        },
    }

# --------------------------------------------------------------------------- #
# Main pipeline                                                               #
# --------------------------------------------------------------------------- #

def safe_yahoo(symbol: str, period1: int = P1) -> pd.DataFrame | None:
    """Fetch a symbol, falling back to the cached raw JSON when the API fails."""
    try:
        return fetch_yahoo(symbol, period1)
    except Exception as e:
        print(f"  WARNING: {symbol} fetch failed ({e}); trying cached copy ...")
        path = RAW / f"{symbol}.json"
        if path.exists():
            try:
                res = json.loads(path.read_text())["chart"]["result"][0]
                ts = res["timestamp"]
                q = res["indicators"]["quote"][0]
                adj = res["indicators"]["adjclose"][0]["adjclose"]
                dates = [datetime.fromtimestamp(t, timezone.utc).astimezone(SYD).date() for t in ts]
                df = pd.DataFrame({"date": pd.to_datetime(dates),
                                   "close": q["close"],
                                   "adjclose": adj}).dropna()
                df = df.drop_duplicates(subset="date").sort_values("date").reset_index(drop=True)
                print(f"  {symbol}: using cached data ({len(df)} rows, to {df.date.max().date()})")
                return df
            except Exception as e2:
                print(f"  WARNING: cached copy for {symbol} unusable ({e2})")
        return None


def safe_asu():
    """AustralianSuper API, falling back to the cached raw JSON."""
    try:
        return fetch_australiansuper()
    except Exception as e:
        print(f"  WARNING: AustralianSuper API failed ({e}); trying cached copy ...")
        path = RAW / "australiansuper_daily.json"
        if path.exists():
            payload = json.loads(path.read_text())
            frames = {}
            for p in payload["drawing"][0]["products"]:
                rows = [(datetime.strptime(g["day"], "%d %b %Y").date(), float(g["rate"]))
                        for g in p["graph"]]
                s = pd.DataFrame(rows, columns=["date", p["title"]])
                s["date"] = pd.to_datetime(s["date"])
                frames[p["title"]] = s.set_index("date")
            df = frames[list(frames)[0]].join([frames[k] for k in list(frames)[1:]], how="outer").sort_index().ffill()
            print(f"  AustralianSuper: using cached data ({len(df)} rows, to {df.index.max().date()})")
            return df
        raise


def safe_official():
    """Official published table, falling back to the cached xlsx."""
    try:
        return fetch_official_table()
    except Exception as e:
        print(f"  WARNING: official table fetch failed ({e}); trying cached copy ...")
        import openpyxl
        path = RAW / "australiansuper_official.xlsx"
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Superannuation"]
        as_of = ws.cell(row=1, column=3).value
        out = {}
        for row in ws.iter_rows(min_row=9, values_only=True):
            if row[2] in ("Australian Shares", "International Shares"):
                out[row[2]] = {"1Y": row[8] * 100, "3Y": row[10] * 100,
                               "5Y": row[11] * 100, "10Y": row[13] * 100,
                               "inception": str(row[17])[:10]}
        return {"asOf": str(as_of)[:10], "options": out}


def main():
    print("Fetching Yahoo Finance data ...")
    ndq = safe_yahoo("NDQ.AX")
    hndq = safe_yahoo("HNDQ.AX")
    tsla = safe_yahoo("TSLA")
    spcx = safe_yahoo("SPCX", period1=1770000000)
    # Major US tech stocks
    googl = safe_yahoo("GOOGL")  # Alphabet Class A
    goog = safe_yahoo("GOOG")     # Alphabet Class C
    nvda = safe_yahoo("NVDA")
    aapl = safe_yahoo("AAPL")
    nflx = safe_yahoo("NFLX")
    amzn = safe_yahoo("AMZN")
    meta = safe_yahoo("META")
    msft = safe_yahoo("MSFT")

    print("Fetching AustralianSuper daily rates ...")
    asu = safe_asu()

    print("Fetching AustralianSuper official returns table ...")
    official = safe_official()
    print(f"  official table as-of {official['asOf']}")

    # ---------------- 200-week MA analysis (NDQ + HNDQ + TSLA + Tech Stocks) ---------------------- #
    ndq_a = analyze_ma200(ndq)
    hndq_a = analyze_ma200(hndq)
    tsla_a = analyze_ma200(tsla)
    # Major US tech stocks
    googl_a = analyze_ma200(googl)
    goog_a = analyze_ma200(goog)
    nvda_a = analyze_ma200(nvda)
    aapl_a = analyze_ma200(aapl)
    nflx_a = analyze_ma200(nflx)
    amzn_a = analyze_ma200(amzn)
    meta_a = analyze_ma200(meta)
    msft_a = analyze_ma200(msft)

    # ---------------- rebased growth indices -------------------------------- #
    ndq_idx = rebase(ndq.set_index("date")["adjclose"])          # total return
    hndq_idx = rebase(hndq.set_index("date")["adjclose"])
    tsla_idx = rebase(tsla.set_index("date")["adjclose"])

    # AustralianSuper: cumulative % since 26 May 2015 -> index 100 at start
    asu_idx = (1 + asu / 100) * 100
    asu_idx = asu_idx[asu_idx.index >= ndq_idx.index[0]]

    def weekly_series(s: pd.Series, start=None):
        s = s.dropna()
        if start is not None:
            s = s[s.index >= start]
        return s.groupby(s.index.to_period("W-SUN")).tail(1)

    def aligned(s: pd.Series, ref: pd.Series, start) -> list:
        wk = weekly_series(s, start)
        return [r2(wk.asof(d)) if not pd.isna(wk.asof(d)) else None for d in ref.index]

    common_start = ndq_idx.index[0]           # 26 May 2015
    growth_dates = weekly_series(asu_idx["International Shares"], common_start)

    growth = {
        "dates": [d.strftime("%Y-%m-%d") for d in growth_dates.index],
        "NDQ (total return)": aligned(ndq_idx, growth_dates, common_start),
        "TSLA (total return)": aligned(tsla_idx, growth_dates, common_start),
        "AusSuper International Shares": aligned(asu_idx["International Shares"], growth_dates, common_start),
        "AusSuper Australian Shares": aligned(asu_idx["Australian Shares"], growth_dates, common_start),
    }

    # hedged vs unhedged since HNDQ inception (Jul 2020)
    h0 = hndq_idx.index[0]
    hndq_rb = rebase(hndq.set_index("date")["adjclose"])
    ndq_rb_h = rebase(ndq.set_index("date")["adjclose"].loc[lambda s: s.index >= h0])
    asi_rb_h = rebase(asu_idx["International Shares"].loc[lambda s: s.index >= h0])
    hedge_dates = weekly_series(ndq_rb_h)
    hedge = {
        "dates": [d.strftime("%Y-%m-%d") for d in hedge_dates.index],
        "NDQ (unhedged)": aligned(ndq_rb_h, hedge_dates, h0),
        "HNDQ (AUD hedged)": aligned(hndq_rb, hedge_dates, h0),
        "AusSuper International Shares": aligned(asi_rb_h, hedge_dates, h0),
    }

    # ---------------- trailing 1/3/5/10 yr (annualised, from real series) ---- #
    trailing = {}
    series_map = {
        "NDQ.AX": ndq_idx,
        "HNDQ.AX": hndq_idx,
        "TSLA": tsla_idx,
        "GOOGL": rebase(googl.set_index("date")["adjclose"]),
        "GOOG": rebase(goog.set_index("date")["adjclose"]),
        "NVDA": rebase(nvda.set_index("date")["adjclose"]),
        "AAPL": rebase(aapl.set_index("date")["adjclose"]),
        "NFLX": rebase(nflx.set_index("date")["adjclose"]),
        "AMZN": rebase(amzn.set_index("date")["adjclose"]),
        "META": rebase(meta.set_index("date")["adjclose"]),
        "MSFT": rebase(msft.set_index("date")["adjclose"]),
        "AusSuper International Shares": asu_idx["International Shares"],
        "AusSuper Australian Shares": asu_idx["Australian Shares"],
    }
    for name, s in series_map.items():
        trailing[name] = {f"{y}Y": r2(trailing_pa(s, y)) for y in (1, 3, 5, 10)}
        trailing[name]["asOf"] = s.dropna().index[-1].strftime("%Y-%m-%d")
        trailing[name]["sinceInception"] = r2(trailing_pa(s, (s.dropna().index[-1] - s.dropna().index[0]).days / 365.25))

    # SPCX: public only since 12 Jun 2026 — raw since-IPO return, not annualised
    sp = spcx.set_index("date")
    sp_close = sp["close"]
    spcx_block = {
        "dates": [d.strftime("%Y-%m-%d") for d in sp_close.index],
        "close": [r2(v) for v in sp_close],
        "price": r2(sp_close.iloc[-1]),
        "date": sp_close.index[-1].strftime("%Y-%m-%d"),
        "ipoPrice": SPCX_IPO_PRICE,
        "ipoDate": "2026-06-12",
        "ipoValuationB": SPCX_IPO_VALUATION_B,
        "impliedSharesB": r2(SPCX_IMPLIED_SHARES_B),
        "impliedMarketCapB": r2(sp_close.iloc[-1] * SPCX_IMPLIED_SHARES_B),
        "peakClose": r2(sp_close.max()),
        "sinceIpoPct": r2((sp_close.iloc[-1] / SPCX_IPO_PRICE - 1) * 100),
        "sinceDebutPct": r2((sp_close.iloc[-1] / sp_close.iloc[0] - 1) * 100),
        # implied market-cap path ($B) for the valuation chart
        "impliedCap": [[d.strftime("%Y-%m-%d"), r2(v * SPCX_IMPLIED_SHARES_B)]
                       for d, v in sp_close.items()],
    }
    trailing["SPCX"] = {"1Y": None, "3Y": None, "5Y": None, "10Y": None,
                        "asOf": spcx_block["date"], "sinceInception": None,
                        "sinceIpoPct": spcx_block["sinceIpoPct"]}

    # ---------------- SpaceX valuation journey (private + public) ------------ #
    spacex_val = {
        "private": [{"date": d, "valB": v, "event": e, "src": s}
                    for d, v, e, s in SPACEX_VALUATIONS],
        "publicImpliedCap": spcx_block["impliedCap"],  # daily $B since IPO
        "note": ("Private era: press-reported funding-round / tender valuations "
                 "(some early dates approximate to month). Public era: SPCX close × "
                 "~13.33B shares implied by the $1.8T IPO at $135/share."),
    }

    # ---------------- payload ------------------------------------------------- #
    kpis = {
        "ndqPrice": ndq_a["kpi"]["price"],
        "ndqDate": ndq_a["kpi"]["date"],
        "ma200": ndq_a["kpi"]["ma200"],
        "vsMaPct": ndq_a["kpi"]["vsMaPct"],
        "inBuyZone": ndq_a["kpi"]["inBuyZone"],
        "weeksInBuyZone": ndq_a["kpi"]["weeksInBuyZone"],
        "weeksWithMa": ndq_a["kpi"]["weeksWithMa"],
        "buyZoneCount": ndq_a["kpi"]["buyZoneCount"],
        "hndqPrice": r2(hndq.iloc[-1]["close"]),
        "hndqDate": hndq.iloc[-1]["date"].strftime("%Y-%m-%d"),
    }

    payload = {
        "generatedAt": datetime.now(SYD).strftime("%Y-%m-%d %H:%M %Z"),
        "sources": [
            {"name": "Yahoo Finance — NDQ.AX / HNDQ.AX / TSLA / SPCX / GOOGL / GOOG / NVDA / AAPL / NFLX / AMZN / META / MSFT daily prices",
             "asOf": kpis["ndqDate"]},
            {"name": "AustralianSuper daily crediting rates API",
             "asOf": asu.index.max().strftime("%Y-%m-%d")},
            {"name": "AustralianSuper official performance report (adviser Excel)",
             "asOf": official["asOf"]},
            {"name": "SpaceX private valuations — CNBC / CNN / Forbes / Bloomberg / BBC press reports",
             "asOf": "2026-06-11"},
        ],
        "kpis": kpis,
        "ndqWeekly": ndq_a["weekly"],
        "buyZones": ndq_a["zones"],
        "hndqWeekly": hndq_a["weekly"],
        "hndqZones": hndq_a["zones"],
        "hndqKpi": hndq_a["kpi"],
        "tslaWeekly": tsla_a["weekly"],
        "tslaZones": tsla_a["zones"],
        "tslaKpi": tsla_a["kpi"],
        "googlWeekly": googl_a["weekly"],
        "googlZones": googl_a["zones"],
        "googlKpi": googl_a["kpi"],
        "googWeekly": goog_a["weekly"],
        "googZones": goog_a["zones"],
        "googKpi": goog_a["kpi"],
        "nvdaWeekly": nvda_a["weekly"],
        "nvdaZones": nvda_a["zones"],
        "nvdaKpi": nvda_a["kpi"],
        "aaplWeekly": aapl_a["weekly"],
        "aaplZones": aapl_a["zones"],
        "aaplKpi": aapl_a["kpi"],
        "nflxWeekly": nflx_a["weekly"],
        "nflxZones": nflx_a["zones"],
        "nflxKpi": nflx_a["kpi"],
        "amznWeekly": amzn_a["weekly"],
        "amznZones": amzn_a["zones"],
        "amznKpi": amzn_a["kpi"],
        "metaWeekly": meta_a["weekly"],
        "metaZones": meta_a["zones"],
        "metaKpi": meta_a["kpi"],
        "msftWeekly": msft_a["weekly"],
        "msftZones": msft_a["zones"],
        "msftKpi": msft_a["kpi"],
        "spcx": spcx_block,
        "spacexVal": spacex_val,
        "growth": growth,
        "hedge": hedge,
        "trailing": trailing,
        "officialAS": official,
    }

    out = ROOT / "data" / "dashboard.js"
    out.write_text("window.DASHBOARD_DATA = " + json.dumps(payload, separators=(",", ":")) + ";\n")
    print(f"\nWrote {out}  ({out.stat().st_size/1024:.0f} KB)")
    print("TSLA:", json.dumps(tsla_a["kpi"]))
    print("TSLA zones:", json.dumps(tsla_a["zones"]))
    print("SPCX:", json.dumps({k: v for k, v in spcx_block.items() if k not in ("dates", "close", "impliedCap")}))
    print("Trailing TSLA:", trailing["TSLA"])

if __name__ == "__main__":
    main()
